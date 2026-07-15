# backend/routes/upload.py
import csv, json, io, os
from fastapi import APIRouter, UploadFile, File, HTTPException, Header, Query, Cookie
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from typing import Optional
from db import get_conn, get_cursor, execute_sql, insert_history, get_user_id
import boto3
from botocore.client import Config
from routes.auth import _get_username_from_auth_header, _decode_access_token
from pydantic import BaseModel

router = APIRouter(prefix="/upload", tags=["Upload"])

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# List of allowed file extensions
ALLOWED_EXTENSIONS = {".csv", ".json", ".pdf", ".txt", ".md", ".docx", ".xlsx"}

# File upload size limits configuration (default: 25MB)
MAX_FILE_SIZE_MB = int(os.getenv("MAX_FILE_SIZE_MB", "25"))
MAX_UPLOAD_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

# Total user storage quota configuration (default: 100MB)
MAX_USER_STORAGE_MB = int(os.getenv("MAX_USER_STORAGE_MB", "100"))
MAX_USER_STORAGE_BYTES = MAX_USER_STORAGE_MB * 1024 * 1024

# S3 Cloud Storage Configurations
S3_BUCKET = os.getenv("S3_BUCKET")
S3_ACCESS_KEY = os.getenv("S3_ACCESS_KEY")
S3_SECRET_KEY = os.getenv("S3_SECRET_KEY")
S3_ENDPOINT_URL = os.getenv("S3_ENDPOINT_URL")
S3_REGION_NAME = os.getenv("S3_REGION_NAME", "us-east-1")

IS_S3 = S3_BUCKET is not None and S3_ACCESS_KEY is not None and S3_SECRET_KEY is not None
_s3_client = None

def get_s3_client():
    global _s3_client
    if not IS_S3:
        return None
    if _s3_client is None:
        kwargs = {
            "aws_access_key_id": S3_ACCESS_KEY,
            "aws_secret_access_key": S3_SECRET_KEY,
            "region_name": S3_REGION_NAME,
        }
        if S3_ENDPOINT_URL:
            kwargs["endpoint_url"] = S3_ENDPOINT_URL
            kwargs["config"] = Config(signature_version="s3v4")
        _s3_client = boto3.client("s3", **kwargs)
    return _s3_client

@router.post("/")
async def upload_file(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    try:
        # Verify authenticated user uploader
        username = _get_username_from_auth_header(authorization, access_token)
        
        # ── Rate Limiter Check ────────────────────────────────────────────────
        from rate_limit import upload_limiter
        upload_limiter.check_rate_limit(username)

        user_id = get_user_id(username)
        if not user_id:
            raise HTTPException(status_code=401, detail="User session invalid")

        filename = os.path.basename(file.filename)
        if not filename or filename in (".", ".."):
            raise HTTPException(status_code=400, detail="Invalid filename")
        _, ext = os.path.splitext(filename.lower())
        
        if ext not in ALLOWED_EXTENSIONS:
            return JSONResponse(
                content={"error": f"Unsupported file type '{ext}'. Supported: CSV, JSON, PDF, TXT, MD."},
                status_code=400
            )

        # Check current total user storage size
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(cur, "SELECT SUM(size) as total_size FROM uploads WHERE user_id = ?", (user_id,))
            row = cur.fetchone()
            current_total = row["total_size"] if row and row["total_size"] is not None else 0

        if current_total >= MAX_USER_STORAGE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"Your total storage quota limit ({MAX_USER_STORAGE_MB}MB) has been reached. Please delete some files first."
            )

        # Create user-specific folder locally
        user_upload_dir = os.path.join(UPLOAD_DIR, username)
        os.makedirs(user_upload_dir, exist_ok=True)
        file_path = os.path.join(user_upload_dir, filename)

        # Stream-read and write to disk in chunks to avoid memory spikes (DoS Protection)
        size = 0
        chunk_size = 1024 * 1024  # 1MB
        try:
            with open(file_path, "wb") as f:
                while True:
                    chunk = await file.read(chunk_size)
                    if not chunk:
                        break
                    size += len(chunk)
                    if size > MAX_UPLOAD_SIZE_BYTES:
                        raise HTTPException(
                            status_code=400,
                            detail=f"File exceeds maximum size limit of {MAX_FILE_SIZE_MB}MB."
                        )
                    if current_total + size > MAX_USER_STORAGE_BYTES:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Uploading this file exceeds your total storage quota limit of {MAX_USER_STORAGE_MB}MB. Please free up space."
                        )
                    f.write(chunk)
        except Exception as e:
            # Clean up the partial file if it was created
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass
            raise e

        # Upload to S3 if configured using the disk path (avoiding loading in memory)
        s3_key = f"{username}/{filename}"
        if IS_S3:
            s3_client = get_s3_client()
            try:
                s3_client.upload_file(file_path, S3_BUCKET, s3_key)
            except Exception as s3_err:
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except Exception:
                        pass
                raise HTTPException(status_code=500, detail=f"S3 cloud upload failed: {s3_err}")

        # Save to database (attaching the user_id)
        db_filepath = f"s3://{S3_BUCKET}/{s3_key}" if IS_S3 else file_path
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(
                cur,
                "INSERT INTO uploads (filename, filepath, size, user_id) VALUES (?, ?, ?, ?)",
                (filename, db_filepath, size, user_id),
            )
            conn.commit()

        # Log to history
        insert_history(username, "file_upload", f"filename={filename}, size={size} bytes")

        # Dynamic RAG Indexing under the active uploader context
        rag_status = "pending"
        chunks_count = 0
        error_msg = None
        
        # Set active user context dynamically during indexing
        from rag import index_file, active_user_context
        token_ctx = active_user_context.set(username)
        try:
            res = index_file(file_path, filename)
            rag_status = res.get("status", "success")
            chunks_count = res.get("chunks", 0)
        except Exception as e:
            rag_status = "failed"
            error_msg = str(e)
            print(f"RAG Indexing failed for {filename}: {e}")
        finally:
            active_user_context.reset(token_ctx)

        # Get preview based on file type (read minimally from disk to avoid memory overhead)
        data_preview = None
        if ext == ".csv":
            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    preview_lines = []
                    for _ in range(100):
                        line = f.readline()
                        if not line:
                            break
                        preview_lines.append(line)
                    preview_text = "".join(preview_lines)
                reader = csv.DictReader(io.StringIO(preview_text))
                data_preview = [row for _, row in zip(range(5), reader)]
            except:
                pass
        elif ext == ".json":
            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    json_text = f.read(10 * 1024)
                data = json.loads(json_text)
                data_preview = data[:5] if isinstance(data, list) else data
            except:
                try:
                    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                        data_preview = f.read(300)
                    if size > 300:
                        data_preview += "..."
                except:
                    pass
        elif ext in {".txt", ".md"}:
            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    data_preview = f.read(300)
                if size > 300:
                    data_preview += "..."
            except:
                pass
        elif ext == ".pdf":
            data_preview = "PDF Document (Preview not available directly)"

        return {
            "message": f"File '{filename}' uploaded successfully",
            "file_path": db_filepath,
            "data_preview": data_preview,
            "rag_status": rag_status,
            "chunks": chunks_count,
            "error": error_msg
        }
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        return JSONResponse(content={"error": str(e)}, status_code=500)

@router.get("/list")
def list_uploads(
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    try:
        username = _get_username_from_auth_header(authorization, access_token)
        user_id = get_user_id(username)
        if not user_id:
            raise HTTPException(status_code=401, detail="User session invalid")

        with get_conn() as conn:
            cur = get_cursor(conn)
            # Filter DB records by user_id
            execute_sql(cur, "SELECT id, filename, filepath, size, uploaded_at FROM uploads WHERE user_id = ? ORDER BY uploaded_at DESC", (user_id,))
            rows = cur.fetchall()
            uploads = [dict(r) for r in rows]
        
        # Merge with ChromaDB indexed files (filtered by active username)
        from rag import active_user_context, get_indexed_files
        token_ctx = active_user_context.set(username)
        try:
            indexed = {item["filename"]: item["chunks"] for item in get_indexed_files()}
        except Exception as e:
            print(f"Error fetching indexed files for list: {e}")
            indexed = {}
        finally:
            active_user_context.reset(token_ctx)
            
        for item in uploads:
            filename = item["filename"]
            item["rag_indexed"] = filename in indexed
            item["chunks"] = indexed.get(filename, 0)
            
        return {"uploads": uploads}
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/reindex/{filename}")
def reindex_file(
    filename: str,
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    """Re-trigger RAG indexing for a file that shows RAG Pending status."""
    try:
        filename = os.path.basename(filename)
        username = _get_username_from_auth_header(authorization, access_token)
        user_id = get_user_id(username)
        if not user_id:
            raise HTTPException(status_code=401, detail="User session invalid")

        # Verify file ownership
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(cur, "SELECT filepath FROM uploads WHERE filename = ? AND user_id = ?", (filename, user_id))
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=403, detail="Access denied or file not found.")

        # Try local path first, then S3 fallback
        local_path = os.path.join(UPLOAD_DIR, username, filename)
        if not os.path.exists(local_path):
            try:
                if IS_S3:
                    s3_client = get_s3_client()
                    os.makedirs(os.path.dirname(local_path), exist_ok=True)
                    s3_client.download_file(S3_BUCKET, f"{username}/{filename}", local_path)
            except Exception as s3_err:
                raise HTTPException(status_code=404, detail=f"File not found locally or in S3: {s3_err}")

        if not os.path.exists(local_path):
            raise HTTPException(status_code=404, detail=f"File '{filename}' not found in storage.")

        # Re-index under the user's context
        from rag import index_file, active_user_context
        token_ctx = active_user_context.set(username)
        rag_status = "pending"
        chunks_count = 0
        error_msg = None
        try:
            res = index_file(local_path, filename)
            rag_status = res.get("status", "success")
            chunks_count = res.get("chunks", 0)
        except Exception as e:
            rag_status = "failed"
            error_msg = str(e)
            print(f"Re-indexing failed for {filename}: {e}")
        finally:
            active_user_context.reset(token_ctx)

        return {
            "message": f"Re-indexing complete for '{filename}'",
            "rag_status": rag_status,
            "chunks": chunks_count,
            "error": error_msg
        }
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))



@router.delete("/{filename}")
def delete_file(
    filename: str,
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    try:
        filename = os.path.basename(filename)
        username = _get_username_from_auth_header(authorization, access_token)
        user_id = get_user_id(username)
        if not user_id:
            raise HTTPException(status_code=401, detail="User session invalid")

        # 1. Verify file ownership
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(cur, "SELECT filepath FROM uploads WHERE filename = ? AND user_id = ?", (filename, user_id))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=403, detail="Access denied. You do not own this file.")
            
            file_path = row["filepath"]
            
            # 2. Delete database entry first (within transaction)
            execute_sql(cur, "DELETE FROM uploads WHERE filename = ? AND user_id = ?", (filename, user_id))
            conn.commit()

        # 3. Delete physical local file if it exists
        local_path = os.path.join(UPLOAD_DIR, username, filename)
        if os.path.exists(local_path):
            os.remove(local_path)
            
        # 4. Delete from S3 if configured
        if IS_S3:
            s3_client = get_s3_client()
            try:
                s3_client.delete_object(Bucket=S3_BUCKET, Key=f"{username}/{filename}")
            except Exception as e:
                print(f"Failed to delete {filename} from S3: {e}")
        
        # 5. Clean up from ChromaDB under active username context
        from rag import active_user_context, delete_file_index
        token_ctx = active_user_context.set(username)
        try:
            delete_file_index(filename)
        except Exception as e:
            print(f"Failed to delete RAG index for {filename}: {e}")
        finally:
            active_user_context.reset(token_ctx)
            
        # 6. Log to history
        insert_history(username, "file_delete", f"filename={filename}")
        
        return {"message": f"File '{filename}' deleted successfully"}
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/download/{filename}")
def download_file(
    filename: str,
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    try:
        filename = os.path.basename(filename)
        # Extract JWT session token from Cookie, Query parameters or Headers
        token_to_decode = access_token or token
        if not token_to_decode and authorization:
            parts = authorization.split(" ")
            if len(parts) == 2 and parts[0].lower() == "bearer":
                token_to_decode = parts[1]

        if not token_to_decode:
            raise HTTPException(status_code=401, detail="Authentication token required to download files")

        username = _decode_access_token(token_to_decode)
        user_id = get_user_id(username)
        if not user_id:
            raise HTTPException(status_code=401, detail="User session invalid")

        # Verify file ownership from relational records
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(cur, "SELECT filepath FROM uploads WHERE filename = ? AND user_id = ?", (filename, user_id))
            row = cur.fetchone()

        if not row:
            raise HTTPException(status_code=403, detail="Access denied. You do not own this file.")

        db_filepath = row["filepath"]
        local_path = os.path.join(UPLOAD_DIR, username, filename)

        # Log to history
        insert_history(username, "file_download", f"filename={filename}")

        import mimetypes
        mime_type, _ = mimetypes.guess_type(filename)
        if not mime_type:
            mime_type = "application/octet-stream"

        # If stored on S3, stream it directly with inline disposition
        if IS_S3 and db_filepath.startswith("s3://"):
            try:
                s3_client = get_s3_client()
                response = s3_client.get_object(Bucket=S3_BUCKET, Key=f"{username}/{filename}")
                return StreamingResponse(
                    response['Body'].iter_chunks(),
                    media_type=mime_type,
                    headers={"Content-Disposition": f"inline; filename=\"{filename}\""}
                )
            except Exception as e:
                print(f"S3 download failed/missing for {filename}: {e}. Falling back to local disk.")
                if not os.path.exists(local_path):
                    raise HTTPException(status_code=404, detail="File not found in S3 or local storage")

        # Local fallback
        if not os.path.exists(local_path):
            raise HTTPException(status_code=404, detail="File not found")

        # Serve inline by specifying headers and omitting the filename argument
        return FileResponse(
            path=local_path,
            media_type=mime_type,
            headers={"Content-Disposition": f"inline; filename=\"{filename}\""}
        )
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/parse-table/{filename}")
def parse_table_file(
    filename: str,
    sheet_name: Optional[str] = Query(None),
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    try:
        filename = os.path.basename(filename)
        token_to_decode = access_token or token
        if not token_to_decode and authorization:
            parts = authorization.split(" ")
            if len(parts) == 2 and parts[0].lower() == "bearer":
                token_to_decode = parts[1]

        if not token_to_decode:
            raise HTTPException(status_code=401, detail="Authentication token required")

        username = _decode_access_token(token_to_decode)
        user_id = get_user_id(username)
        if not user_id:
            raise HTTPException(status_code=401, detail="User session invalid")

        # Verify file ownership from relational records
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(cur, "SELECT filepath FROM uploads WHERE filename = ? AND user_id = ?", (filename, user_id))
            row = cur.fetchone()

        if not row:
            raise HTTPException(status_code=403, detail="Access denied. You do not own this file.")

        db_filepath = row["filepath"]
        local_path = os.path.join(UPLOAD_DIR, username, filename)

        # Download from S3 if missing locally
        if IS_S3 and db_filepath.startswith("s3://") and not os.path.exists(local_path):
            try:
                s3_client = get_s3_client()
                user_folder = os.path.join(UPLOAD_DIR, username)
                os.makedirs(user_folder, exist_ok=True)
                s3_client.download_file(S3_BUCKET, f"{username}/{filename}", local_path)
            except Exception as e:
                raise HTTPException(status_code=404, detail=f"File not found on cloud storage: {e}")

        if not os.path.exists(local_path):
            raise HTTPException(status_code=404, detail="File not found")

        _, ext = os.path.splitext(filename.lower())
        
        sheet_names = []
        active_sheet = ""
        headers = []
        rows = []

        if ext == ".csv":
            # Read CSV
            try:
                content = ""
                # Try UTF-8 encoding
                try:
                    with open(local_path, "r", encoding="utf-8") as f:
                        content = f.read()
                except UnicodeDecodeError:
                    # Fallback to latin-1
                    with open(local_path, "r", encoding="latin-1") as f:
                        content = f.read()
                
                f_io = io.StringIO(content)
                reader = csv.reader(f_io)
                all_rows = list(reader)
                if all_rows:
                    headers = [h.strip() for h in all_rows[0]]
                    for r in all_rows[1:]:
                        rows.append([val.strip() for val in r])
            except Exception as csv_err:
                raise HTTPException(status_code=400, detail=f"Failed to parse CSV file: {csv_err}")

        elif ext == ".xlsx":
            # Read XLSX
            try:
                import openpyxl
                wb = openpyxl.load_workbook(local_path, data_only=True)
                sheet_names = wb.sheetnames
                
                # Determine active sheet
                if sheet_name and sheet_name in sheet_names:
                    active_sheet = sheet_name
                else:
                    active_sheet = sheet_names[0] if sheet_names else ""
                
                if active_sheet:
                    sheet = wb[active_sheet]
                    raw_rows = []
                    for row in sheet.iter_rows(values_only=True):
                        # Filter out fully empty rows
                        if not any(val is not None for val in row):
                            continue
                        raw_rows.append([str(val).strip() if val is not None else "" for val in row])
                    
                    if raw_rows:
                        headers = raw_rows[0]
                        rows = raw_rows[1:]
            except Exception as xlsx_err:
                raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {xlsx_err}")
        else:
            raise HTTPException(status_code=400, detail="Only CSV and XLSX formats can be parsed as grids.")

        return {
            "sheet_names": sheet_names,
            "active_sheet": active_sheet,
            "headers": headers,
            "rows": rows
        }

    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))





class EditFileRequest(BaseModel):
    content: str


@router.put("/edit/{filename}")
def edit_file(
    filename: str,
    req: EditFileRequest,
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None)
):
    try:
        filename = os.path.basename(filename)
        # Verify authenticated user
        username = _get_username_from_auth_header(authorization, access_token)
        user_id = get_user_id(username)
        if not user_id:
            raise HTTPException(status_code=401, detail="User session invalid")

        # 1. Verify file ownership (Database read)
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(cur, "SELECT filepath, size FROM uploads WHERE filename = ? AND user_id = ?", (filename, user_id))
            row = cur.fetchone()

        if not row:
            raise HTTPException(status_code=403, detail="Access denied. You do not own this file.")

        old_size = row["size"] if row and row["size"] is not None else 0

        # Check total user storage usage
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(cur, "SELECT SUM(size) as total_size FROM uploads WHERE user_id = ?", (user_id,))
            sum_row = cur.fetchone()
            current_total = sum_row["total_size"] if sum_row and sum_row["total_size"] is not None else 0

        content_bytes = req.content.encode("utf-8")
        new_size = len(content_bytes)

        if current_total - old_size + new_size > MAX_USER_STORAGE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"Updating this file would exceed your total storage quota limit of {MAX_USER_STORAGE_MB}MB. Please delete other files first."
            )

        # 2. Check if file is an editable text-based file
        _, ext = os.path.splitext(filename.lower())
        if ext not in [".txt", ".md", ".json", ".csv"]:
            raise HTTPException(status_code=400, detail="Only text-based files (.txt, .md, .json, .csv) can be edited.")

        local_path = os.path.join(UPLOAD_DIR, username, filename)

        # 3. Clean up previous RAG embeddings first
        from rag import active_user_context, delete_file_index, index_file
        token_ctx = active_user_context.set(username)
        try:
            delete_file_index(filename)
        except Exception as e:
            print(f"Failed to clear RAG index for edit: {e}")

        # 4. Overwrite physical file
        try:
            with open(local_path, "wb") as f:
                f.write(content_bytes)
        except Exception as e:
            active_user_context.reset(token_ctx)
            raise HTTPException(status_code=500, detail=f"Failed to write file update: {e}")

        # 5. S3 update if configured
        if IS_S3:
            s3_client = get_s3_client()
            try:
                s3_client.upload_file(local_path, S3_BUCKET, f"{username}/{filename}")
            except Exception as s3_err:
                active_user_context.reset(token_ctx)
                raise HTTPException(status_code=500, detail=f"S3 cloud update failed: {s3_err}")

        # 6. Re-index file in RAG
        chunks_count = 0
        try:
            res = index_file(local_path, filename)
            chunks_count = res.get("chunks", 0)
        except Exception as e:
            print(f"RAG Re-indexing failed during edit: {e}")
        finally:
            active_user_context.reset(token_ctx)

        # 7. Update database record with new size (Database write)
        with get_conn() as conn:
            cur = get_cursor(conn)
            execute_sql(cur, "UPDATE uploads SET size = ? WHERE filename = ? AND user_id = ?", (new_size, filename, user_id))
            conn.commit()

        # 8. Log to history
        insert_history(username, "file_edit", f"filename={filename}, new_size={new_size}")

        return {
            "message": f"File '{filename}' updated successfully",
            "size": new_size,
            "chunks": chunks_count
        }

    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

